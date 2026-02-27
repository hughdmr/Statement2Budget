import pandas as pd
import os
import argparse
import json
from datetime import datetime
from copy import copy
from call_llm import call_llm


def load_categories(filepath: str = None) -> tuple[list, list, list]:
    """Charge les catégories et les exclusions depuis categories.json."""
    if filepath is None:
        filepath = os.path.join(os.path.dirname(__file__), 'categories.json')
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data['depenses'], data['revenus'], data.get('exclusions', [])


# Charger catégories et exclusions depuis `categories.json`
CATEGORIES_DEPENSES, CATEGORIES_REVENUS, EXCLUDE_PATTERNS = load_categories()

def load_releve(filepath: str) -> pd.DataFrame:
    """Charge un relevé bancaire SG au format CSV."""
    # Le fichier a un header spécial sur la première ligne, on skip
    df = pd.read_csv(
        filepath,
        sep=';',
        skiprows=2,
        encoding='latin-1',
        decimal=','
    )
    df.columns = ['Date', 'Libelle', 'Detail', 'Montant', 'Devise']
    
    # Filtrer les virements internes et paris sportifs (patterns définis dans categories.json)
    mask = df['Detail'].str.upper().apply(
        lambda d: not any(p in d for p in EXCLUDE_PATTERNS)
    )
    df = df[mask].reset_index(drop=True)
    
    return df


def load_template(filepath: str) -> dict:
    """Charge le template CSV et retourne deux DataFrames (dépenses et revenus)."""
    # Le template a 2 tableaux côte à côte: Dépenses (colonnes 1-4) et Revenus (colonnes 6-9)
    # Lignes 1-3: instruction/titres/vide, Ligne 4: en-têtes colonnes
    df = pd.read_csv(filepath, encoding='utf-8', header=None, skiprows=3)
    
    # Extraire les colonnes pour dépenses (indices 1-4) et revenus (indices 6-9)
    # La première ligne est l'en-tête (Date, Montant...), on la saute
    depenses = df.iloc[1:, 1:5].copy()
    depenses.columns = ['Date', 'Montant', 'Description', 'Catégorie']
    depenses = depenses.dropna(how='all').reset_index(drop=True)
    
    revenus = df.iloc[1:, 6:10].copy()
    revenus.columns = ['Date', 'Montant', 'Description', 'Catégorie']
    revenus = revenus.dropna(how='all').reset_index(drop=True)
    
    return {'depenses': depenses, 'revenus': revenus}


def load_exemples(releves_dir: str, budgets_dir: str) -> dict:
    """Charge les exemples de relevés et templates remplis."""
    exemples = {}
    print(f"Chargement des exemples depuis {releves_dir} et {budgets_dir}...")
    
    if not os.path.exists(releves_dir) or not os.path.exists(budgets_dir):
        return exemples
    
    # Lister les relevés disponibles
    releves = [f for f in os.listdir(releves_dir) if f.endswith('.csv')]
    budgets = [f for f in os.listdir(budgets_dir) if f.endswith('.csv')]
    
    for releve_file in releves:
        # Extraire le mois du nom de fichier (ex: janvier2026.csv -> janvier)
        month_key = releve_file.replace('.csv', '').rstrip('0123456789')
        year = releve_file.replace('.csv', '')[len(month_key):]
        
        # Trouver le budget correspondant (Budget Janvier 2026 -> janvier)
        matching_budget = None
        for budget_file in budgets:
            if month_key in budget_file.lower() and year in budget_file:
                matching_budget = budget_file
                break
        
        if matching_budget:
            exemples[month_key] = {
                'releve': load_releve(os.path.join(releves_dir, releve_file)),
                'template': load_template(os.path.join(budgets_dir, matching_budget))
            }
    
    return exemples


def build_prompt(releve: pd.DataFrame, template: dict, exemples: dict = None) -> str:
    """Construit le prompt pour l'assistant comptable."""
    
    categories_str = f"""
Catégories pour les DÉPENSES (uniquement pour les dépenses) : {', '.join(CATEGORIES_DEPENSES)}
Catégories pour les REVENUS (uniquement pour les revenus) : {', '.join(CATEGORIES_REVENUS)}
ATTENTION : les catégories de dépenses et de revenus sont complètement indépendantes. Ne mélange jamais les catégories entre les deux sections.
"""
    
    releve_str = releve[['Date', 'Detail', 'Montant']].to_csv(index=False, sep=';')
    template_depenses = template['depenses'].to_csv(index=False, sep=';')
    template_revenus = template['revenus'].to_csv(index=False, sep=';')
    
    prompt = f"""Tu es un assistant comptable. En entrée tu prendras un csv avec l'export de mes transactions bancaires mensuelles. Je veux que tu les tries et complète mon template de budget csv.

Voici les catégories disponibles :
{categories_str}

"""
    
    # Ajouter les exemples s'il y en a
    if exemples:
        prompt += "Voici des exemples de relevés et templates remplis :\n"
        for month, data in exemples.items():
            if 'releve' in data and 'template' in data:
                ex_releve = data['releve'][['Date', 'Detail', 'Montant']].to_csv(index=False, sep=';')
                ex_depenses = data['template']['depenses'].to_csv(index=False, sep=';')
                ex_revenus = data['template']['revenus'].to_csv(index=False, sep=';')
                prompt += f"\n--- Exemple {month} ---\n"
                prompt += f"Relevé:\n{ex_releve}\n"
                prompt += f"Template rempli - DÉPENSES:\n{ex_depenses}\n"
                prompt += f"Template rempli - REVENUS:\n{ex_revenus}\n"
    
    prompt += f"""
Voici le template vide à remplir :

--- DÉPENSES ---
{template_depenses}

--- REVENUS ---
{template_revenus}

Voici le relevé bancaire à traiter :
{releve_str}

Instructions :
Prend exemple sur janvier et février pour remplir le template du relevé donné. Ne change pas les descriptions, dates ou montants. Juste trie les transactions à ajouter avec les catégories correspondantes en te basant sur les derniers relevés.

IMPORTANT : Retourne UNIQUEMENT le CSV brut avec des points-virgules (;) comme séparateur, sans aucune phrase, sans explication, sans markdown, sans ```csv```, sans commentaire. Juste les deux sections DÉPENSES et REVENUS au format CSV séparées par une ligne vide.
Format attendu exact :
--- DÉPENSES ---
Date;Montant;Description;Catégorie
...lignes...

--- REVENUS ---
Date;Montant;Description;Catégorie
...lignes...
"""
    
    return prompt


def format_montant(val) -> str:
    """Formate un montant en 'X,XX €' (valeur absolue, virgule décimale, symbole €)."""
    try:
        n = abs(float(str(val).replace(',', '.')))
    except (ValueError, TypeError):
        return val
    # Ex: 1,234.50 € -> 1234,50 €
    formatted = f"{n:.2f}".replace('.', ',')
    return f"{formatted} €"


def parse_llm_output(result: str) -> tuple[list[list[str]], list[list[str]]]:
    """Parse la réponse brute du LLM en deux listes de lignes [Date, Montant, Description, Catégorie]."""
    depenses = []
    revenus = []

    # Trouver les deux sections
    lines = result.strip().splitlines()
    current_section = None
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if 'DÉPENSES' in stripped.upper() or 'DEPENSES' in stripped.upper():
            current_section = 'depenses'
            continue
        if 'REVENUS' in stripped.upper():
            current_section = 'revenus'
            continue
        # Ignorer les lignes d'en-tête CSV
        if stripped.lower().startswith('date'):
            continue

        # Déterminer le séparateur (point-virgule ou virgule)
        if ';' in stripped:
            parts = stripped.split(';')
        else:
            # Fallback: virgule — mais attention aux montants "X,XX"
            parts = stripped.split(',')

        if len(parts) >= 4:
            date, montant, description, categorie = parts[0], parts[1], parts[2], parts[3]
            row = [date.strip(), montant.strip(), description.strip(), categorie.strip()]
            if current_section == 'depenses':
                depenses.append(row)
            elif current_section == 'revenus':
                revenus.append(row)

    return depenses, revenus


def save_result(result: str, output_path: str):
    """Parse la réponse LLM et sauvegarde au format template (deux tableaux côte à côte)."""
    depenses, revenus = parse_llm_output(result)

    # Construire les lignes du CSV final
    lines = []
    # Ligne 1 : instruction
    lines.append(',"Pour modifier ou ajouter des catégories, modifiez les tableaux ""Dépenses"" et ""Revenus"" de la feuille ""Récapitulatif"".",,,,,,,,')
    # Ligne 2 : titres des sections
    lines.append(',Dépenses,,,,,Revenus,,,')
    # Ligne 3 : vide
    lines.append(',,,,,,,,,')
    # Ligne 4 : en-têtes colonnes
    lines.append(',Date,Montant,Description,Catégorie,,Date,Montant,Description,Catégorie')

    max_rows = max(len(depenses), len(revenus))
    for i in range(max_rows):
        # Colonnes dépenses (1-4)
        if i < len(depenses):
            d = depenses[i]
            d_date = d[0]
            d_montant = format_montant(d[1])
            d_desc = d[2]
            d_cat = d[3]
        else:
            d_date = d_montant = d_desc = d_cat = ''

        # Colonnes revenus (6-9)
        if i < len(revenus):
            r = revenus[i]
            r_date = r[0]
            r_montant = format_montant(r[1])
            r_desc = r[2]
            r_cat = r[3]
        else:
            r_date = r_montant = r_desc = r_cat = ''

        # Quoting des montants (contiennent une virgule)
        if d_montant:
            d_montant = f'"{d_montant}"'
        if r_montant:
            r_montant = f'"{r_montant}"'

        line = f',{d_date},{d_montant},{d_desc},{d_cat},,{r_date},{r_montant},{r_desc},{r_cat}'
        lines.append(line)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')

    print(f"Résultat sauvegardé dans : {output_path}")


def save_to_xlsx(result: str, template_xlsx: str, output_xlsx: str):
    """Parse la réponse LLM et remplit la feuille Transactions du template xlsx."""
    import openpyxl

    depenses, revenus = parse_llm_output(result)

    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb['Transactions']

    # Récupérer le style d'une ligne de données existante (ligne 5) pour le copier
    def get_cell_style(cell):
        """Copie le format d'une cellule."""
        return {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'alignment': copy(cell.alignment),
        }

    # Sauvegarder les styles de la ligne 5 (première ligne de données)
    styles = {}
    for col_letter in ['B', 'C', 'D', 'E', 'G', 'H', 'I', 'J']:
        cell = ws[f'{col_letter}5']
        styles[col_letter] = get_cell_style(cell)

    # Effacer les données existantes (à partir de la ligne 5)
    for row in range(5, ws.max_row + 1):
        for col_letter in ['B', 'C', 'D', 'E', 'G', 'H', 'I', 'J']:
            ws[f'{col_letter}{row}'].value = None

    def parse_date(date_str: str):
        """Parse une date dd/mm/yyyy en datetime."""
        try:
            return datetime.strptime(date_str.strip(), '%d/%m/%Y')
        except (ValueError, AttributeError):
            return date_str

    def parse_montant(val_str: str) -> float:
        """Parse un montant string en float (valeur absolue)."""
        try:
            return abs(float(str(val_str).replace(',', '.')))
        except (ValueError, TypeError):
            return 0.0

    # Écrire les dépenses (colonnes B-E, à partir de la ligne 5)
    for i, row in enumerate(depenses):
        r = 5 + i
        ws[f'B{r}'] = parse_date(row[0])
        ws[f'C{r}'] = parse_montant(row[1])
        ws[f'D{r}'] = row[2]
        ws[f'E{r}'] = row[3]
        # Appliquer les styles
        for col_letter in ['B', 'C', 'D', 'E']:
            for attr, val in styles[col_letter].items():
                setattr(ws[f'{col_letter}{r}'], attr, val)

    # Écrire les revenus (colonnes G-J, à partir de la ligne 5)
    for i, row in enumerate(revenus):
        r = 5 + i
        ws[f'G{r}'] = parse_date(row[0])
        ws[f'H{r}'] = parse_montant(row[1])
        ws[f'I{r}'] = row[2]
        ws[f'J{r}'] = row[3]
        # Appliquer les styles
        for col_letter in ['G', 'H', 'I', 'J']:
            for attr, val in styles[col_letter].items():
                setattr(ws[f'{col_letter}{r}'], attr, val)

    wb.save(output_xlsx)
    print(f"Résultat xlsx sauvegardé dans : {output_xlsx}")


def main():
    parser = argparse.ArgumentParser(description='Assistant budget - Catégorise les transactions bancaires')
    parser.add_argument('--releve', nargs='?', help='Chemin vers le fichier de relevé bancaire CSV')
    parser.add_argument('--output', '-o', help='Chemin de sortie pour le template rempli')
    parser.add_argument('--template', '-t', default='data/template/template_feuille_transactions.csv', help='Chemin vers le template')
    
    args = parser.parse_args()
    
    # Chemins par défaut
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, args.template)
    
    # Si pas de relevé spécifié, utiliser le dernier dans sog_releve
    if args.releve:
        releve_path = args.releve
    else:
        releve_dir = os.path.join(base_dir, 'data', 'inputs', 'sog_releve')
        releves = sorted([f for f in os.listdir(releve_dir) if f.endswith('.csv')])
        if not releves:
            print("Erreur: Aucun relevé trouvé dans data/inputs/sog_releve/")
            return
        releve_path = os.path.join(releve_dir, releves[-1])
        print(f"Utilisation du relevé: {releves[-1]}")
    
    # Chemin de sortie
    if args.output:
        output_path = args.output
    else:
        releve_name = os.path.splitext(os.path.basename(releve_path))[0]
        output_path = os.path.join(base_dir, 'data', 'outputs', f'budget_{releve_name}.csv')
    
    # Charger les données
    print("Chargement des données...")
    releve = load_releve(releve_path)
    print(f"Relevé chargé: {releve.head(3)}")
    print(f"Relevé chargé: {len(releve)} transactions")
    template = load_template(template_path)
    print(f"Template chargé: {template['depenses'].head(3)}")
    
    releves_dir = os.path.join(base_dir, 'data', 'inputs', 'sog_releve')
    budgets_dir = os.path.join(base_dir, 'data', 'inputs', 'exemples_budget')
    exemples = load_exemples(releves_dir, budgets_dir)
    print(f"Exemples chargés: {exemples.keys()}")
    
    # Construire le prompt
    prompt = build_prompt(releve, template, exemples)
    
    # Sauvegarder le prompt pour debug
    prompt_path = os.path.join(base_dir, 'data', 'outputs', 'debug_prompt.txt')
    with open(prompt_path, 'w', encoding='utf-8') as f:
        f.write(prompt)
    print(f"Prompt sauvegardé dans : {prompt_path}")
    
    # Appeler le LLM
    print("Appel à un model...")
    result = call_llm(prompt)
    
    # Sauvegarder le résultat
    save_result(result, output_path)
    
    # Sauvegarder dans le xlsx
    template_xlsx = os.path.join(base_dir, 'data', 'template', 'template_a_remplir.xlsx')
    releve_name = os.path.splitext(os.path.basename(releve_path))[0]
    output_xlsx = os.path.join(base_dir, 'data', 'outputs', f'budget_{releve_name}.xlsx')
    save_to_xlsx(result, template_xlsx, output_xlsx)
    
    print("Terminé!")


if __name__ == "__main__":
    main()